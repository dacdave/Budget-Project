from BudgetCalculations import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def create_cash_flow(start, end, account, initial=0):
    wb = Workbook()
    dest_filename = "Test_File.xlsx"
    ws = wb.active
    ws.title = "Cash Flow"
    
    ws['A1'] = "Date"
    ws['B1'] = "Description"
    ws['C1'] = "Amount"
    ws['D1'] = "Account Balance"
    
    ws['A2'] = start
    ws['B2'] = "Initial Balance"
    ws['C2'] = initial
    ws['D2'] = initial

    data = account.ledger.coalate_transactions(start, end, account)
    for i in data:
        a = 'A' + str(3 + data.index(i))
        b = 'B' + str(3 + data.index(i))
        c = 'C' + str(3 + data.index(i))
        d = 'D' + str(3 + data.index(i))
        ws[a] = i.date
        ws[b] = i.description
        ws[c] = i.amount
        if data.index(i) == 0:
            balance = initial + i.amount
        else:
            balance += i.amount
        ws[d] = balance
    wb.save(filename = dest_filename)

def category_spending(start, end, account, initial=0):
    pass

